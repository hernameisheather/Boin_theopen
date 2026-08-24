"""
학생 성적/피드백 공유 포털
- 선생님: 관리자 페이지에서 학생/기록을 직접 추가·수정·삭제 (Excel 자동 저장)
- 학부모: 학생코드 + PIN으로 본인 자녀의 기록만 열람
"""
import os
import io
import random
from datetime import datetime, timedelta
from collections import defaultdict
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file
)
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# ─── 환경 변수 ────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

APP_SECRET = os.environ.get("APP_SECRET", "change-me-in-production")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "teacher1234")
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(os.path.dirname(__file__), "data"))
EXCEL_PATH = os.path.join(DATA_DIR, "students.xlsx")

os.makedirs(DATA_DIR, exist_ok=True)

# ─── Flask 앱 ────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = APP_SECRET
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB

# ─── 데이터 ──────────────────────────────────────────────────
_data_cache = {
    "mtime": 0,
    "students": {},
    "records": [],
    "homeworks": [],      # 이번주 숙제 목록: [{"content": "...", "published": True/False}]
    "messages": [],       # 신쌤의 한마디: [{"student_code": "...", "content": "...", "published": True/False}]
}


def _parse_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if v is None:
        return ""
    return str(v).strip()


def _parse_workbook(wb):
    """openpyxl Workbook을 (students, records, homework, messages)로 파싱.
    `기록` 시트는 두 가지 형식 모두 자동 감지:
      - 신형식: 날짜 | 학생코드 | 학생이름 | 항목 | 점수 | 피드백
      - 구형식: 날짜 | 학생코드 | 항목 | 점수 | 피드백
    `공지사항` 시트(선택): 종류 | 대상학생코드 | 내용
    """
    students = {}
    if "학생명단" in wb.sheetnames:
        ws = wb["학생명단"]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or row[0] is None:
                continue
            code = str(row[0]).strip()
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            pin = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            parent = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            if code:
                students[code] = {"name": name, "pin": pin, "parent": parent}

    records = []
    if "기록" in wb.sheetnames:
        ws = wb["기록"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            has_name_col = len(header) >= 3 and header[2] in ("학생이름", "이름", "name", "Name")

            # 비고(플래그)와 완료(처리됨) 컬럼 위치 자동 탐지 — 어디에 있어도 OK
            flag_idx = None
            resolved_idx = None
            term_idx = None
            for i, h in enumerate(header):
                if h in ("비고", "특이사항", "플래그", "flag", "메모"):
                    flag_idx = i
                if h in ("완료", "해결", "처리", "처리됨", "resolved", "done"):
                    resolved_idx = i
                if h in ("학기/과정", "학기", "과정", "term", "session"):
                    term_idx = i

            def _parse_resolved(v):
                if v is None:
                    return False
                s = str(v).strip().lower()
                return s in ("o", "완료", "true", "1", "y", "예", "✓", "v", "처리")

            for row in rows[1:]:
                if not row or row[0] is None or (len(row) > 1 and row[1] is None):
                    continue
                if has_name_col:
                    rec = {
                        "date": _parse_date(row[0]),
                        "student_code": str(row[1]).strip(),
                        "category": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                        "score": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                        "feedback": str(row[5]).strip() if len(row) > 5 and row[5] else "",
                    }
                else:
                    rec = {
                        "date": _parse_date(row[0]),
                        "student_code": str(row[1]).strip(),
                        "category": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                        "score": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                        "feedback": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                    }
                rec["flag"] = (str(row[flag_idx]).strip()
                               if flag_idx is not None and len(row) > flag_idx and row[flag_idx]
                               else "")
                rec["resolved"] = (_parse_resolved(row[resolved_idx])
                                   if resolved_idx is not None and len(row) > resolved_idx
                                   else False)
                rec["term"] = (str(row[term_idx]).strip()
                               if term_idx is not None and len(row) > term_idx and row[term_idx]
                               else "")
                records.append(rec)

    homeworks = []
    messages = []
    if "공지사항" in wb.sheetnames:
        ws = wb["공지사항"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            has_name_col = len(header) >= 3 and header[2] in ("학생이름", "이름", "name", "Name")
            content_idx = 3 if has_name_col else 2

            # 상태(게시/내림) 컬럼 자동 탐지
            status_idx = None
            for i, h in enumerate(header):
                if h in ("상태", "게시", "공개", "status", "published"):
                    status_idx = i

            def _is_published(v):
                if v is None:
                    return True  # 기본값은 게시
                s = str(v).strip().lower()
                if not s:
                    return True
                # 내림/비공개로 처리할 값들
                return s not in ("내림", "비공개", "off", "no", "n", "0", "false", "unpublished", "비활성", "x")

            for row in rows[1:]:
                if not row or row[0] is None:
                    continue
                kind = str(row[0]).strip()
                code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                content = str(row[content_idx]).strip() if len(row) > content_idx and row[content_idx] else ""
                if not content:
                    continue
                published = True
                if status_idx is not None and len(row) > status_idx:
                    published = _is_published(row[status_idx])

                if kind in ("이번주숙제", "이번주 숙제", "숙제", "homework"):
                    homeworks.append({"content": content, "published": published})
                elif kind in ("한마디", "신쌤의한마디", "신쌤의 한마디", "메시지", "message"):
                    messages.append({"student_code": code, "content": content, "published": published})

    return students, records, homeworks, messages


def load_data():
    """Excel을 읽어서 메모리에 캐싱. 파일 수정시각이 바뀌면 재로딩."""
    if not os.path.exists(EXCEL_PATH):
        _data_cache.update({"mtime": 0, "students": {}, "records": [], "homeworks": [], "messages": []})
        return _data_cache

    mtime = os.path.getmtime(EXCEL_PATH)
    if mtime == _data_cache["mtime"]:
        return _data_cache

    wb = load_workbook(EXCEL_PATH, data_only=True)
    students, records, homeworks, messages = _parse_workbook(wb)

    _data_cache.update({
        "mtime": mtime,
        "students": students,
        "records": records,
        "homeworks": homeworks,
        "messages": messages,
    })
    return _data_cache


def _add_records_dropdowns(ws):
    """`기록` 시트의 비고(G열) / 완료(H열)에 드롭다운 추가."""
    # 비고 컬럼 — 재시/숙제미비/추가과제/결석/우수
    dv_flag = DataValidation(
        type="list",
        formula1='"재시,숙제미비,추가과제,결석,우수"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv_flag.error = "재시 / 숙제미비 / 추가과제 / 결석 / 우수 중에서 선택하세요. (자유 입력이 필요하면 데이터 검증 해제 후 입력)"
    dv_flag.errorTitle = "잘못된 비고"
    dv_flag.prompt = "드롭다운에서 선택"
    dv_flag.promptTitle = "비고 선택"
    ws.add_data_validation(dv_flag)
    dv_flag.add("G2:G2000")

    # 완료 컬럼 — O / 완료
    dv_done = DataValidation(
        type="list",
        formula1='"O,완료"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv_done.error = "처리됐으면 O 또는 완료를, 미처리는 비워두세요."
    dv_done.errorTitle = "잘못된 완료 표시"
    dv_done.prompt = "처리됐으면 O 선택"
    dv_done.promptTitle = "완료 표시"
    ws.add_data_validation(dv_done)
    dv_done.add("H2:H2000")


def _add_term_dropdown(ws):
    """`기록` 시트의 학기/과정(I열)에 드롭다운 추가."""
    options = ",".join(TERMS)
    dv = DataValidation(
        type="list",
        formula1=f'"{options}"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.error = f"{' 또는 '.join(TERMS)} 중에서 선택하세요."
    dv.errorTitle = "잘못된 학기"
    ws.add_data_validation(dv)
    dv.add("I2:I2000")


def _add_announcements_dropdown(ws):
    """`공지사항` 시트의 상태(E열)에 게시/내림 드롭다운."""
    dv = DataValidation(
        type="list",
        formula1='"게시,내림"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.error = "게시 또는 내림 중에서 선택하세요."
    dv.errorTitle = "잘못된 상태"
    ws.add_data_validation(dv)
    dv.add("E2:E1000")


def save_data(students=None, records=None, homeworks=None, messages=None):
    """현재 메모리 상태를 Excel 파일로 저장.
    None을 전달하면 현재 캐시 값을 그대로 유지 (덮어쓰기 안 함).
    """
    if students is None or records is None or homeworks is None or messages is None:
        current = load_data()
        if students is None:
            students = current["students"]
        if records is None:
            records = current["records"]
        if homeworks is None:
            homeworks = current["homeworks"]
        if messages is None:
            messages = current["messages"]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "학생명단"
    ws1.append(["학생코드", "학생이름", "PIN", "학부모이름(선택)"])
    for code, s in students.items():
        ws1.append([code, s.get("name", ""), s.get("pin", ""), s.get("parent", "")])

    ws2 = wb.create_sheet("기록")
    ws2.append(["날짜", "학생코드", "학생이름", "항목", "점수", "피드백", "비고", "완료", "학기/과정"])
    for r in records:
        code = r.get("student_code", "")
        name = students.get(code, {}).get("name", "")
        ws2.append([
            r.get("date", ""),
            code,
            name,
            r.get("category", ""),
            r.get("score", ""),
            r.get("feedback", ""),
            r.get("flag", ""),
            "O" if r.get("resolved") else "",
            r.get("term", ""),
        ])

    ws3 = wb.create_sheet("공지사항")
    ws3.append(["종류", "대상학생코드", "학생이름", "내용", "상태"])
    for h in homeworks:
        ws3.append([
            "이번주숙제",
            "",
            "",
            h.get("content", ""),
            "게시" if h.get("published") else "내림",
        ])
    for m in messages:
        code = m.get("student_code", "")
        name = students.get(code, {}).get("name", "") if code else ""
        ws3.append([
            "한마디",
            code,
            name,
            m.get("content", ""),
            "게시" if m.get("published") else "내림",
        ])

    for ws, widths in [
        (ws1, [10, 12, 8, 18]),
        (ws2, [12, 10, 12, 14, 8, 40, 12, 8, 14]),
        (ws3, [14, 14, 12, 50, 10]),
    ]:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    _add_records_dropdowns(ws2)
    _add_term_dropdown(ws2)
    _add_announcements_dropdown(ws3)

    wb.save(EXCEL_PATH)
    _data_cache["mtime"] = 0  # 캐시 무효화
    load_data()


def _parse_uploaded_excel(file_storage):
    """업로드된 Excel 파일을 임시 저장 후 파싱.
    반환: (students, records, homework, messages)
    """
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        file_storage.save(tmp.name)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        return _parse_workbook(wb)
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def records_for(student_code):
    """특정 학생의 기록을 날짜 내림차순으로 반환, 카테고리별 그룹핑."""
    data = load_data()
    items = [r for r in data["records"] if r["student_code"] == student_code]
    items.sort(key=lambda r: r["date"], reverse=True)

    grouped = defaultdict(list)
    for r in items:
        grouped[r["category"] or "기타"].append(r)
    return items, dict(grouped)


# ─── 학기/과정 (Term) 구분 ────────────────────────────────
TERMS = ["2학기 정규반", "방학특강"]
DEFAULT_TERM = "2학기 정규반"
_TERM_MIGRATION_MARKER = os.path.join(DATA_DIR, ".term_migration_v1.json")


# ─── 단어 시험 회차 정의 ──────────────────────────────────
WORD_TEST_ROUNDS = [
    {"label": "1회차", "range": "31-34"},
    {"label": "2회차", "range": "35-38"},
    {"label": "3회차", "range": "39-42"},
    {"label": "4회차", "range": "43-45"},
]


def _record_matches_round(record, range_str):
    """기록이 특정 단어시험 회차에 해당하는지 판단.
    매칭 기준: 항목 또는 피드백에 range 문자열이 들어있고, 단어/어휘 키워드도 함께 있을 때.
    """
    cat = (record.get("category") or "").strip()
    fb = (record.get("feedback") or "").strip()
    if not cat and not fb:
        return False
    # 공백 정규화
    range_norm = range_str.replace(" ", "")
    cat_norm = cat.replace(" ", "")
    fb_norm = fb.replace(" ", "")
    # 항목이 정확히 회차 범위인 경우 (예: "16-20")
    if cat_norm == range_norm:
        return True
    # 항목/피드백에 range 포함 + 단어/어휘 키워드 존재
    has_range = (range_norm in cat_norm) or (range_norm in fb_norm)
    if not has_range:
        return False
    keywords = ["단어", "어휘", "vocab", "word"]
    return any(k in (cat + fb).lower() if k in ("vocab", "word") else k in (cat + fb)
               for k in keywords)


def _is_word_test_completed(score):
    """점수 텍스트가 '실제로 응시 완료'인지 판단.
    '미완료', '미응시', '미제출', '결석' 등이 포함되면 False.
    """
    if not score:
        return False
    s = str(score).strip()
    if not s:
        return False
    incomplete_markers = ["미완료", "미응시", "미시행", "미제출", "결석", "패스"]
    for marker in incomplete_markers:
        if marker in s:
            return False
    return True


# ─── Review test (항목명 기반) ───────────────────────────────
def compute_review_test_records(student_records):
    """Review test 관련 기록을 시험명(항목명)별로 그룹핑.
    매칭: 항목명에 'review' 또는 '리뷰' 포함.
    같은 항목명에 여러 기록이 있으면 가장 최근 날짜 기록만 사용.
    """
    by_cat = {}
    for r in student_records:
        cat = (r.get("category") or "").strip()
        if not cat:
            continue
        cat_lower = cat.lower()
        if "review" not in cat_lower and "리뷰" not in cat_lower:
            continue
        date = r.get("date") or ""
        if cat not in by_cat or date > (by_cat[cat].get("date") or ""):
            by_cat[cat] = r

    result = []
    for cat, r in by_cat.items():
        score_text = r.get("score", "")
        numeric_score = _parse_score(score_text)
        completed = _is_word_test_completed(score_text)
        result.append({
            "test_name": cat,
            "has_record": True,
            "completed": completed,
            "score": score_text,
            "numeric_score": numeric_score,
            "date": r.get("date") or "",
            "feedback": r.get("feedback") or "",
        })
    # 날짜 내림차순(최신이 위에) 정렬
    result.sort(key=lambda x: (x["date"] or ""), reverse=True)
    return result


# ─── 평가 시험 (4개: 중간평가 1·2차, 파이널 1·2차) ────────────
EVALUATION_TESTS = [
    {"label": "보신반 재시험 (D6-9)", "key": "retake_boin", "is_retake": False},
    {"label": "중간평가 1차", "key": "midterm_1", "is_retake": False},
    {"label": "중간평가 2차 (재시)", "key": "midterm_2", "is_retake": True,
     "retake_threshold": 38},
    {"label": "파이널 1차", "key": "final_1", "is_retake": False},
    {"label": "파이널 2차", "key": "final_2", "is_retake": False},
]


def _record_matches_evaluation(record, key):
    """기록이 특정 평가에 해당하는지 — 항목명으로 판별.
    규칙:
      - '중간평가'(또는 'midterm') 키워드 있음 → 중간평가
      - 그게 아니고 '파이널'/'final'/'기말' 키워드 있음 → 파이널
      - '2차'/'재시' 명시 있으면 → 2차, 없으면 기본적으로 1차로 분류
      - '기말대비 중간평가'는 중간평가가 우선
    """
    cat = (record.get("category") or "").strip()
    if not cat:
        return False
    cat_norm = cat.replace(" ", "").lower()

    # 보신반 재시험: 항목명 정확 일치만 (다른 평가 키와 충돌 없음)
    if key == "retake_boin":
        return cat_norm == RETAKE_EXAM_CATEGORY.replace(" ", "").lower()

    # 중간평가 우선 판단
    is_midterm = ("중간평가" in cat_norm or "midterm" in cat_norm)
    if not is_midterm:
        is_midterm = "중간" in cat_norm and "기말대비" not in cat_norm

    # 파이널/기말은 중간평가가 아닐 때만
    is_final = False
    if not is_midterm:
        is_final = ("파이널" in cat_norm or "final" in cat_norm or "기말" in cat_norm)

    # 2차 명시 표시 (없으면 1차로 기본 분류)
    is_2nd = ("2차" in cat_norm) or ("2회" in cat_norm) or ("재시" in cat_norm)

    if key == "midterm_1":
        return is_midterm and not is_2nd
    if key == "midterm_2":
        return is_midterm and is_2nd
    if key == "final_1":
        return is_final and not is_2nd
    if key == "final_2":
        return is_final and is_2nd
    return False


def compute_evaluation_status(student_records):
    """평가 4개 상태. 중간평가 2차는 1차 ≤38인 경우만 응시 대상."""
    # 1차 점수 먼저 산출
    mid1_matches = sorted(
        [r for r in student_records if _record_matches_evaluation(r, "midterm_1")],
        key=lambda r: r.get("date", ""), reverse=True
    )
    mid1_latest = mid1_matches[0] if mid1_matches else None
    mid1_score = _parse_score(mid1_latest.get("score", "")) if mid1_latest else None

    result = []
    for ev in EVALUATION_TESTS:
        key = ev["key"]
        matching = sorted(
            [r for r in student_records if _record_matches_evaluation(r, key)],
            key=lambda r: r.get("date", ""), reverse=True
        )
        latest = matching[0] if matching else None
        score_text = latest.get("score", "") if latest else None
        numeric_score = _parse_score(score_text) if score_text else None
        has_record = latest is not None
        completed = has_record and _is_word_test_completed(score_text)

        # 재시험(중간평가 2차) 대상 여부 판단
        is_eligible = True
        eligibility_note = None
        if ev.get("is_retake"):
            threshold = ev.get("retake_threshold", 38)
            if mid1_score is None:
                is_eligible = False
                eligibility_note = "중간평가 1차 미응시"
            elif mid1_score > threshold:
                is_eligible = False
                eligibility_note = f"재시험 대상 아님 (1차 {int(mid1_score) if mid1_score == int(mid1_score) else mid1_score}점)"
            else:
                eligibility_note = f"재시험 대상 (1차 {int(mid1_score) if mid1_score == int(mid1_score) else mid1_score}점 ≤ {threshold})"

        result.append({
            "label": ev["label"],
            "key": key,
            "is_retake": ev.get("is_retake", False),
            "has_record": has_record,
            "completed": completed,
            "score": score_text,
            "numeric_score": numeric_score,
            "date": latest.get("date") if latest else None,
            "feedback": latest.get("feedback") if latest else None,
            "is_eligible": is_eligible,
            "eligibility_note": eligibility_note,
        })
    return result


# ─── 보신반 재시험 (Day 6-9) — 인포털 응시 + 서버 채점 ─────────
RETAKE_EXAM_CATEGORY = "보신반 재시험 (D6-9)"
RETAKE_EXAM_MAX = 50
RETAKE_MC_POINTS = 3
# 결과를 구글 시트(Apps Script 웹앱)로도 전송 — 비우면 전송 안 함
RETAKE_GSHEET_ENDPOINT = os.environ.get(
    "RETAKE_GSHEET_ENDPOINT",
    "https://script.google.com/macros/s/AKfycby6ZylRrCS5r0PsHHbf4pBlij3i8bbqCEUDFk-sqhtD7EOLc2C5ICWFdR5XmELaXrMV/exec",
)
_RETAKE_CIRCLED = {0: "-", 1: "①", 2: "②", 3: "③", 4: "④", 5: "⑤"}


def _forward_to_gsheet(payload):
    """결과를 구글 Apps Script 시트로 best-effort 전송 (실패해도 응시에는 영향 없음)."""
    if not RETAKE_GSHEET_ENDPOINT:
        return
    import json as _json
    import urllib.request as _urlreq
    try:
        body = _json.dumps(payload).encode("utf-8")
        req = _urlreq.Request(RETAKE_GSHEET_ENDPOINT, data=body,
                              headers={"Content-Type": "application/json"})
        _urlreq.urlopen(req, timeout=10)
    except Exception as e:
        print(f"[WARN] gsheet forward failed: {e}")

# 객관식 정답(1-indexed) + 해설 — exam.html 정답키와 동일
RETAKE_MC_KEY = [
    {"n": 1, "ans": 1, "exp": "단순 지표는 중요한 목표에 '가까이' 데려갈 수 없다는 흐름. ①closer → no closer/further."},
    {"n": 2, "ans": 1, "exp": "바로 뒤 'The survivors'는 선택이 일부를 제거했다는 앞 내용이 필요. ①."},
    {"n": 3, "ans": 2, "exp": "청소년의 변화한 자율성 욕구에 맞게 부모 방식이 조정되어야 한다. ②."},
    {"n": 4, "ans": 4, "exp": "도덕적 위험 제기 → (B) 취약 사용자 기만 → (A) 만화처럼 제안 → (C) 부연. ④ (B)-(A)-(C)."},
    {"n": 5, "ans": 5, "exp": "be동사 보어 자리이므로 부사 unconsciously → 형용사 unconscious. ⑤."},
    {"n": 6, "ans": 4, "exp": "인과를 '설명/규명'하려는 시도가 오히려 흐리게 만든다. ④obscure는 부적절."},
    {"n": 7, "ans": 5, "exp": "표본추출의 정보 손실을 넘어 전체 데이터를 활용하자는 글. ⑤."},
    {"n": 8, "ans": 3, "exp": "긍정적 공상은 에너지를 '빼앗는다'는 흐름. ③boost는 부적절."},
    {"n": 9, "ans": 2, "exp": "유럽 통치자의 무능 vs. 고도로 중앙집권화된 아시아/중동. ②."},
    {"n": 10, "ans": 3, "exp": "'Popular thinking said ~, yet 인물 ~' 예시의 마지막 항목. ③."},
    {"n": 11, "ans": 5, "exp": "개도국은 추가 비용 흡수가 '더 어렵다'고 했으므로 '손쉽게 적응'은 불일치. ⑤."},
]

RETAKE_SA1_ANSWER = ["Shared", "thinking", "tools", "allow", "students", "to",
                     "link", "different", "academic", "disciplines"]
RETAKE_SA1_MODEL = "Shared thinking tools allow students to link different academic disciplines."
RETAKE_SA1_EXP = "공유된 사고 도구(shared thinking tools)가 학생들이 서로 다른 학문 분야를 연결하게 해 준다는 요지."
RETAKE_SA2_A = "relaxing"
RETAKE_SA2_B = "acquiring"
RETAKE_SA2_MODEL = "(A) relaxing  (B) acquiring"
RETAKE_SA2_EXP = "benefits from + 동명사(relaxing), thereby + 동명사(acquiring)."
RETAKE_SA3_ANSWER = ["searching", "for", "causal", "explanations", "only", "creates", "more", "confusion"]
RETAKE_SA3_MODEL = "searching for causal explanations only creates more confusion"
RETAKE_SA3_EXP = "명확한 상관관계라도 인과를 설명하려 들면 오히려 더 혼란스러워진다는 요지."


def _norm_ans(s):
    """exam.html의 norm과 동일 규칙: 소문자 + 따옴표/마침표/쉼표 제거 + 공백 단일화."""
    s = (s or "").lower().replace('"', '').replace('.', '').replace(',', '')
    return " ".join(s.split())


def grade_retake(answers):
    """클라이언트 답안(dict)을 서버에서 채점.
    answers = {"mc": {"1": 3, ...}, "sa1": [...], "sa2": {"A": "..", "B": ".."}, "sa3": [...]}
    """
    answers = answers or {}
    mc_in = answers.get("mc", {}) or {}

    score = 0
    mc_right = 0
    mc_results = []
    for q in RETAKE_MC_KEY:
        try:
            pick = int(mc_in.get(str(q["n"]), 0) or 0)
        except (ValueError, TypeError):
            pick = 0
        ok = pick == q["ans"]
        if ok:
            score += RETAKE_MC_POINTS
            mc_right += 1
        mc_results.append({"n": q["n"], "pick": pick, "ans": q["ans"], "ok": ok, "exp": q["exp"]})

    sa_results = []
    got1 = 5 if _norm_ans(" ".join(answers.get("sa1", []))) == _norm_ans(" ".join(RETAKE_SA1_ANSWER)) else 0
    sa_results.append({"n": "서술형 1", "got": got1, "max": 5, "ok": got1 == 5,
                       "model": RETAKE_SA1_MODEL, "exp": RETAKE_SA1_EXP})

    sa2 = answers.get("sa2", {}) or {}
    a_ok = _norm_ans(sa2.get("A", "")) == RETAKE_SA2_A
    b_ok = _norm_ans(sa2.get("B", "")) == RETAKE_SA2_B
    got2 = (4 if a_ok else 0) + (4 if b_ok else 0)
    sa_results.append({"n": "서술형 2", "got": got2, "max": 8, "ok": got2 == 8,
                       "a_ok": a_ok, "b_ok": b_ok, "model": RETAKE_SA2_MODEL, "exp": RETAKE_SA2_EXP})

    got3 = 4 if _norm_ans(" ".join(answers.get("sa3", []))) == _norm_ans(" ".join(RETAKE_SA3_ANSWER)) else 0
    sa_results.append({"n": "서술형 3", "got": got3, "max": 4, "ok": got3 == 4,
                       "model": RETAKE_SA3_MODEL, "exp": RETAKE_SA3_EXP})

    sa_score = got1 + got2 + got3
    score += sa_score
    mc_str = f"{mc_right}/{len(RETAKE_MC_KEY)}"
    sa_str = f"{sa_score}/17"
    feedback = (f"객관식 {mc_str} · 서술형 {sa_str} "
                f"(서술형1 {got1}/5·서술형2 {got2}/8·서술형3 {got3}/4)")
    mc_detail = " | ".join(f"{m['n']}:{_RETAKE_CIRCLED.get(m['pick'], '-')}" for m in mc_results)
    sa_detail = " | ".join(f"{s['n']}:{s['got']}/{s['max']}" for s in sa_results)
    detail = f"{mc_detail} | {sa_detail}"
    return {
        "score": score, "max": RETAKE_EXAM_MAX,
        "mc_str": mc_str, "sa_str": sa_str, "feedback": feedback, "detail": detail,
        "mc_results": mc_results, "sa_results": sa_results,
    }


def compute_word_test_status(student_records):
    """특정 학생의 기록에서 단어시험 회차별 상태 계산.
    같은 회차에 여러 기록이 있으면 가장 최근 날짜 기록을 사용.
    상태 분류:
      - completed: 기록 있음 + 점수가 '미완료' 등이 아닌 경우
      - has_record: 기록이 존재 (점수 무관)
    """
    result = []
    for round_info in WORD_TEST_ROUNDS:
        matching = [r for r in student_records
                    if _record_matches_round(r, round_info["range"])]
        matching.sort(key=lambda r: r.get("date", ""), reverse=True)
        latest = matching[0] if matching else None
        score_text = latest.get("score", "") if latest else None
        numeric_score = _parse_score(score_text) if score_text else None
        has_record = latest is not None
        completed = has_record and _is_word_test_completed(score_text)
        result.append({
            "label": round_info["label"],
            "range": round_info["range"],
            "has_record": has_record,
            "completed": completed,
            "score": score_text,
            "numeric_score": numeric_score,
            "date": latest.get("date") if latest else None,
            "feedback": latest.get("feedback") if latest else None,
            "attempts": len(matching),
        })
    return result


# ─── 클리닉 자동 기록 (매주 화/목) ────────────────────────────
CLINIC_CATEGORY = "주간 클리닉"
CLINIC_BASE_FEEDBACK = "주간 혼공학습지 풀이 및 서술형 피드백 완료."
CLINIC_TIPS = [
    "[학습 포인트] 지문 속 복잡한 구문 — 주절과 종속절을 먼저 분리해서 분석하는 연습 필요.",
    "[학습 포인트] 직역이 안되는 영어 표현 — 문맥 속 의역으로 접근하는 훈련 강조.",
    "[학습 포인트] 3줄 도식화를 통한 전체 주제 파악 — 도입/전개/결론 구조 의식.",
    "[학습 포인트] 주제 배열 영작 서술형 — 한국어 해석이 없는 주제 영작 시 핵심 키워드 먼저 정리.",
    "[학습 포인트] 요약문 영작 — 원문의 군더더기를 빼고 핵심 메시지만 간결하게 표현.",
    "[학습 포인트] 지문 속 추상명사 — 구체적 예시로 의미를 풀어 이해하는 습관.",
    "[학습 포인트] 관계대명사절 — 선행사부터 다시 짚어가며 정확히 해석.",
    "[학습 포인트] 도치·강조 구문 — 평서문으로 재배열 후 의미 확인.",
]
# 요일별 정기 클리닉 명단 — 코드에 하드코딩하지 않고 파일에 저장(관리자 화면에서 편집).
# 자동 업로드(스케줄러)는 제거됨. 이 명단은 "이 명단으로 한 번에 추가" 편의 기능에만 사용.
CLINIC_ROSTER_FILE = os.path.join(DATA_DIR, "clinic_roster.json")
CLINIC_WEEKDAYS = [(1, "화요일"), (3, "목요일"), (5, "토요일")]
_CLINIC_ROSTER_SEED = {
    "1": ["정주원", "유한선"],
    "3": ["박서원", "장우영", "오우진", "이성우"],
    "5": ["박도환"],
}


def load_clinic_roster():
    """요일(str)→학생이름 리스트. 파일 없으면 시드값으로 생성."""
    import json
    if os.path.exists(CLINIC_ROSTER_FILE):
        try:
            with open(CLINIC_ROSTER_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {str(k): list(v) for k, v in data.items()}
        except Exception as e:
            print(f"[WARN] clinic roster load failed: {e}")
    return {k: list(v) for k, v in _CLINIC_ROSTER_SEED.items()}


def save_clinic_roster(roster):
    import json
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(CLINIC_ROSTER_FILE, "w", encoding="utf-8") as f:
        json.dump({str(k): list(v) for k, v in roster.items()}, f, ensure_ascii=False, indent=2)


def _find_or_create_clinic_student(name, students):
    """이름으로 학생을 찾고, 없으면 CLI### 코드로 신규 생성."""
    for code, s in students.items():
        if s.get("name") == name:
            return code
    nums = []
    for c in students:
        if c.startswith("CLI"):
            try:
                nums.append(int(c[3:]))
            except ValueError:
                pass
    new_code = f"CLI{(max(nums) + 1 if nums else 1):03d}"
    students[new_code] = {"name": name, "pin": "1234", "parent": ""}
    return new_code


def _resolve_names_to_codes(names):
    """학생 이름 리스트를 코드 리스트로. 없는 이름은 CLI 코드로 생성(저장)."""
    data = load_data()
    students = dict(data["students"])
    codes = []
    created = False
    before = set(students.keys())
    for n in names:
        n = (n or "").strip()
        if not n:
            continue
        code = _find_or_create_clinic_student(n, students)
        codes.append(code)
    if set(students.keys()) != before:
        created = True
    if created:
        save_data(students, list(data["records"]))
    return codes


def add_clinic_records(clinic_date_str, student_codes, auto_feedback=True, note="", per_code_note=None):
    """클리닉 기록을 코드 기준으로 추가/업데이트.
    - auto_feedback=True: 표준 문구 + 학습 포인트 자동 생성
    - auto_feedback=False: note(직접 입력) 를 피드백으로 사용
    - 같은 날짜·학생 기록이 이미 있으면: note 있으면 [추가]로 누적, 없으면 건너뜀
    반환: (added, updated, skipped) 학생명 리스트
    """
    per_code_note = per_code_note or {}
    data = load_data()
    students = dict(data["students"])
    records = list(data["records"])

    existing_idx = {}
    for i, r in enumerate(records):
        if (r.get("date") == clinic_date_str
                and r.get("category") == CLINIC_CATEGORY):
            existing_idx[r.get("student_code")] = i

    added_names, updated_names, skipped_names = [], [], []

    for code in student_codes:
        name = students.get(code, {}).get("name", code)
        extra = (per_code_note.get(code, "") or note or "").strip()

        if code in existing_idx:
            if extra:
                idx = existing_idx[code]
                cur_fb = records[idx].get("feedback", "")
                line = f"[추가] {extra}"
                if line not in cur_fb:
                    records[idx]["feedback"] = cur_fb.rstrip() + ("\n" if cur_fb else "") + line
                    updated_names.append(name)
                else:
                    skipped_names.append(name)
            else:
                skipped_names.append(name)
        else:
            if auto_feedback:
                feedback = f"{CLINIC_BASE_FEEDBACK} {random.choice(CLINIC_TIPS)}"
                if extra:
                    feedback += f"\n[추가] {extra}"
            else:
                feedback = extra or "주간 클리닉 진행."
            records.append({
                "date": clinic_date_str,
                "student_code": code,
                "category": CLINIC_CATEGORY,
                "score": "완료",
                "feedback": feedback,
                "flag": "",
                "resolved": False,
                "term": DEFAULT_TERM,
            })
            added_names.append(name)

    if added_names or updated_names:
        save_data(students, records)
    return added_names, updated_names, skipped_names


def _this_week_weekday_date(weekday_num):
    """이번주 (월~일 기준) weekday(0=Mon, 1=Tue, 3=Thu)의 날짜 문자열 반환."""
    today = datetime.now()
    diff = weekday_num - today.weekday()
    target = today + timedelta(days=diff)
    return target.strftime("%Y-%m-%d")


def _last_weekday_date(weekday_num):
    """오늘 또는 가장 가까운 과거 weekday의 날짜 문자열 반환."""
    today = datetime.now()
    diff = (today.weekday() - weekday_num) % 7
    target = today - timedelta(days=diff)
    return target.strftime("%Y-%m-%d")


# 자동 클리닉 스케줄러/부트스트랩 제거됨 (완전 수동 전환).
# 클리닉 기록은 관리자 화면(/admin/clinic)에서 직접 추가/삭제한다.


# ─── 인증 ────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return wrapper


def parent_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("student_code"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


# ─── 라우트: 학부모 ──────────────────────────────────────────
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        code = request.form.get("student_code", "").strip()
        pin = request.form.get("pin", "").strip()

        data = load_data()
        student = data["students"].get(code)
        if not student:
            flash("학생코드가 올바르지 않습니다.", "error")
        elif student["pin"] and student["pin"] != pin:
            flash("PIN 번호가 일치하지 않습니다.", "error")
        else:
            session["student_code"] = code
            session["student_name"] = student["name"]
            return redirect(url_for("my_page"))

    return render_template("login.html")


def _parse_score(raw):
    """문자열 점수에서 숫자만 추출 (예: '92점' → 92.0). 실패 시 None."""
    if raw is None:
        return None
    s = str(raw).replace("점", "").replace(" ", "")
    try:
        return float(s)
    except (ValueError, AttributeError):
        return None


def _build_chart(points, width=560, height=160, padding=32):
    """날짜순으로 정렬된 [{date, score}] 리스트를 SVG 좌표 데이터로 변환."""
    if not points:
        return None
    inner_w = width - 2 * padding
    inner_h = height - 2 * padding
    n = len(points)
    scores = [p["score"] for p in points]
    min_s, max_s = min(scores), max(scores)
    rng = max(max_s - min_s, 10)
    y_min = max(0, min_s - rng * 0.15)
    y_max = min(100, max_s + rng * 0.15) if max_s <= 100 else max_s + rng * 0.15
    if y_max - y_min < 10:
        y_max = y_min + 10

    def x_at(i):
        return width / 2 if n == 1 else padding + (i / (n - 1)) * inner_w

    def y_at(v):
        return padding + (1 - (v - y_min) / (y_max - y_min)) * inner_h

    plot = [{
        "x": round(x_at(i), 1),
        "y": round(y_at(p["score"]), 1),
        "date": p["date"],
        "score": p["score"],
    } for i, p in enumerate(points)]

    path = "M " + " L ".join(f"{p['x']},{p['y']}" for p in plot)

    return {
        "width": width,
        "height": height,
        "padding": padding,
        "y_min": int(round(y_min)),
        "y_max": int(round(y_max)),
        "points": plot,
        "path": path,
    }


@app.route("/me")
@parent_required
def my_page():
    code = session["student_code"]
    data = load_data()
    student = data["students"].get(code)
    if not student:
        session.clear()
        return redirect(url_for("login"))

    # 학부모/학생 페이지는 항상 2학기 정규반만 표시 (방학특강은 관리자 페이지에서만 조회 가능)
    selected_term = DEFAULT_TERM

    def _term_of(r):
        return r.get("term") or DEFAULT_TERM

    # 이 학생의 2학기 정규반 기록만
    items = [r for r in data["records"]
             if r["student_code"] == code and _term_of(r) == selected_term]
    items.sort(key=lambda r: r.get("date", ""), reverse=True)

    grouped = defaultdict(list)
    for r in items:
        grouped[r["category"] or "기타"].append(r)
    grouped = dict(grouped)

    # 학기 필터가 적용된 전체 기록 (반 평균/등수 계산용)
    term_all_records = [r for r in data["records"] if _term_of(r) == selected_term]

    # 학생 본인 카테고리별 통계
    stats = {}
    for cat, recs in grouped.items():
        nums = [v for v in (_parse_score(r["score"]) for r in recs) if v is not None]
        if nums:
            stats[cat] = {
                "count": len(nums),
                "avg": round(sum(nums) / len(nums), 1),
                "max": max(nums),
                "min": min(nums),
            }

    # 반 전체 카테고리별 평균 (해당 학기 학생들만)
    class_buckets = defaultdict(list)
    for r in term_all_records:
        v = _parse_score(r["score"])
        if v is not None:
            class_buckets[r["category"] or "기타"].append(v)
    class_avg = {
        cat: round(sum(vs) / len(vs), 1)
        for cat, vs in class_buckets.items() if vs
    }

    # 카테고리별 추세 그래프 (이 학생, 이 학기만)
    charts = {}
    for cat, recs in grouped.items():
        pts = []
        for r in recs:
            v = _parse_score(r["score"])
            if v is not None and r["date"]:
                pts.append({"date": r["date"], "score": v})
        pts.sort(key=lambda p: p["date"])
        if len(pts) >= 2:
            charts[cat] = _build_chart(pts)

    # 이 학생에게 보여줄 한마디: 게시 중 + (공통 OR 본인 개별)
    my_messages = [
        m for m in data["messages"]
        if m.get("published", True)
        and (not m.get("student_code") or m.get("student_code") == code)
    ]

    # 게시 중인 숙제만 (2학기 정규반에서만 표시)
    if selected_term == DEFAULT_TERM:
        published_homeworks = [h["content"] for h in data["homeworks"] if h.get("published", True)]
        homework = "\n\n".join(published_homeworks) if published_homeworks else ""
    else:
        homework = ""

    # 반 등수 (항목명 기준, 학기 필터 적용)
    cat_best = defaultdict(dict)
    for r in term_all_records:
        sc = _parse_score(r["score"])
        if sc is None:
            continue
        cat = r["category"] or "기타"
        sc_code = r["student_code"]
        if sc_code not in cat_best[cat] or sc > cat_best[cat][sc_code]:
            cat_best[cat][sc_code] = sc

    ranks = {}
    for cat, code_scores in cat_best.items():
        if len(code_scores) < 2:
            continue
        sorted_entries = sorted(code_scores.items(), key=lambda x: -x[1])
        prev_score = None
        current_rank = 0
        for i, (c, s) in enumerate(sorted_entries):
            if s != prev_score:
                current_rank = i + 1
                prev_score = s
            ranks[f"{c}|{cat}"] = (current_rank, len(code_scores))

    # 날짜별 그룹핑
    by_date_dict = defaultdict(list)
    for r in items:
        by_date_dict[r["date"]].append(r)
    by_date = sorted(by_date_dict.items(), key=lambda x: x[0], reverse=True)

    # 미처리 특이사항
    my_flags = [r for r in items if r.get("flag") and not r.get("resolved")]
    my_flags.sort(key=lambda r: r["date"], reverse=True)

    # zone 계산 (모두 학기 필터된 items 기준)
    word_tests = compute_word_test_status(items)
    word_tests_done = sum(1 for t in word_tests if t["completed"])
    word_tests_total = len(word_tests)

    review_records = compute_review_test_records(items)
    review_done = sum(1 for r in review_records if r["completed"])
    review_total = len(review_records)

    return render_template(
        "student.html",
        student=student,
        records=items,
        grouped=grouped,
        stats=stats,
        class_avg=class_avg,
        charts=charts,
        homework=homework,
        messages=my_messages,
        by_date=by_date,
        ranks=ranks,
        my_flags=my_flags,
        word_tests=word_tests,
        word_tests_done=word_tests_done,
        word_tests_total=word_tests_total,
        review_records=review_records,
        review_done=review_done,
        review_total=review_total,
        last_update=datetime.fromtimestamp(data["mtime"]).strftime("%Y-%m-%d %H:%M") if data["mtime"] else "—",
    )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─── 라우트: 보신반 재시험 (학생 응시) ───────────────────────
@app.route("/retake")
@parent_required
def retake_exam():
    code = session["student_code"]
    data = load_data()
    student = data["students"].get(code)
    if not student:
        session.clear()
        return redirect(url_for("login"))

    # 이전 응시 기록(있으면 마지막 점수 표시)
    prev_score = None
    for r in data["records"]:
        if (r.get("student_code") == code
                and r.get("category") == RETAKE_EXAM_CATEGORY):
            prev_score = r.get("score")

    return render_template(
        "retake_exam.html",
        student_name=student["name"],
        prev_score=prev_score,
        exam_max=RETAKE_EXAM_MAX,
    )


@app.route("/retake/submit", methods=["POST"])
@parent_required
def retake_submit():
    code = session["student_code"]
    data = load_data()
    if code not in data["students"]:
        return {"error": "unknown student"}, 400

    answers = request.get_json(silent=True) or {}
    result = grade_retake(answers)

    # 같은 학생의 기존 보신반 재시험 기록 제거 후 1건으로 교체
    today = datetime.now().strftime("%Y-%m-%d")
    records = [r for r in data["records"]
               if not (r.get("student_code") == code
                       and r.get("category") == RETAKE_EXAM_CATEGORY)]
    records.append({
        "date": today,
        "student_code": code,
        "category": RETAKE_EXAM_CATEGORY,
        "score": str(result["score"]),
        "feedback": result["feedback"],
        "flag": "",
        "resolved": False,
    })
    save_data(data["students"], records)

    # 구글 시트로도 전송 (백그라운드 — 실패해도 응시·저장에는 영향 없음)
    if RETAKE_GSHEET_ENDPOINT:
        import threading
        stamp = (datetime.utcnow() + timedelta(hours=9)).strftime("%Y. %m. %d. %H:%M:%S")
        payload = {
            "exam": "보신반 재시험 D6-D9",
            "name": data["students"][code]["name"],
            "score": result["score"], "max": result["max"],
            "mc": result["mc_str"], "sa": result["sa_str"],
            "time": stamp, "detail": result["detail"],
        }
        threading.Thread(target=_forward_to_gsheet, args=(payload,), daemon=True).start()

    return result


@app.route("/admin/retake/backfill-google")
@admin_required
def retake_backfill_google():
    """이미 저장된 보신반 재시험 기록을 구글 시트로 일괄 전송 (한 번만 사용)."""
    data = load_data()
    sent = 0
    for r in data["records"]:
        if r.get("category") != RETAKE_EXAM_CATEGORY:
            continue
        name = data["students"].get(r.get("student_code"), {}).get("name", r.get("student_code", ""))
        _forward_to_gsheet({
            "exam": "보신반 재시험 D6-D9 (기존기록)",
            "name": name,
            "score": r.get("score", ""), "max": RETAKE_EXAM_MAX,
            "mc": "", "sa": "",
            "time": r.get("date", ""),
            "detail": r.get("feedback", ""),
        })
        sent += 1
    flash(f"기존 재시험 결과 {sent}건을 구글 시트로 전송했습니다. (중복 방지를 위해 한 번만 실행하세요)", "success")
    return redirect(url_for("admin"))


# ─── 라우트: 관리자 ──────────────────────────────────────────
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin"))
        flash("관리자 비밀번호가 올바르지 않습니다.", "error")
    return render_template("admin_login.html")


@app.route("/admin", methods=["GET", "POST"])
@admin_required
def admin():
    if request.method == "POST" and "excel_file" in request.files:
        f = request.files["excel_file"]
        mode = request.form.get("mode", "overwrite")

        if f.filename:
            filename = secure_filename(f.filename)
            if not filename.lower().endswith((".xlsx", ".xlsm")):
                flash("Excel(.xlsx) 파일만 업로드 가능합니다.", "error")
            else:
                try:
                    new_students, new_records, new_homeworks, new_messages = _parse_uploaded_excel(f)
                except Exception as e:
                    flash(f"엑셀 파일을 읽을 수 없습니다: {e}", "error")
                    return redirect(url_for("admin"))

                # 업로드 시 학기/과정이 비어있으면 자동으로 기본값(2학기 정규반) 지정
                defaulted_term_count = 0
                for r in new_records:
                    if not r.get("term"):
                        r["term"] = DEFAULT_TERM
                        defaulted_term_count += 1

                if mode == "append":
                    current = load_data()

                    # 학생: 신규 코드만 추가, 기존 코드는 건너뜀
                    merged_students = dict(current["students"])
                    added_students = 0
                    skipped_students = 0
                    for code, s in new_students.items():
                        if code not in merged_students:
                            merged_students[code] = s
                            added_students += 1
                        else:
                            skipped_students += 1

                    # 기록 중복 판별
                    def _rec_key(r):
                        return (r.get("date", ""), r.get("student_code", ""),
                                r.get("category", ""), r.get("score", ""),
                                r.get("feedback", ""))
                    seen_recs = {_rec_key(r) for r in current["records"]}
                    unique_records = []
                    skipped_records = 0
                    for r in new_records:
                        k = _rec_key(r)
                        if k in seen_recs:
                            skipped_records += 1
                        else:
                            seen_recs.add(k)
                            unique_records.append(r)
                    merged_records = current["records"] + unique_records

                    # 숙제 중복 판별: content 동일
                    seen_hws = {h.get("content", "") for h in current["homeworks"]}
                    unique_hws = []
                    skipped_hws = 0
                    for h in new_homeworks:
                        if h.get("content", "") in seen_hws:
                            skipped_hws += 1
                        else:
                            seen_hws.add(h.get("content", ""))
                            unique_hws.append(h)
                    merged_homeworks = current["homeworks"] + unique_hws

                    # 한마디 중복 판별: (학생코드, 내용) 일치
                    def _msg_key(m):
                        return (m.get("student_code", ""), m.get("content", ""))
                    seen_msgs = {_msg_key(m) for m in current["messages"]}
                    unique_msgs = []
                    skipped_msgs = 0
                    for m in new_messages:
                        k = _msg_key(m)
                        if k in seen_msgs:
                            skipped_msgs += 1
                        else:
                            seen_msgs.add(k)
                            unique_msgs.append(m)
                    merged_messages = current["messages"] + unique_msgs

                    save_data(merged_students, merged_records, merged_homeworks, merged_messages)

                    parts = []
                    parts.append(f"신규 학생 {added_students}명" +
                                 (f" (중복 {skipped_students}명 건너뜀)" if skipped_students else ""))
                    parts.append(f"새 기록 {len(unique_records)}건" +
                                 (f" (중복 {skipped_records}건 건너뜀)" if skipped_records else ""))
                    parts.append(f"새 숙제 {len(unique_hws)}건" +
                                 (f" (중복 {skipped_hws}건 건너뜀)" if skipped_hws else ""))
                    parts.append(f"새 한마디 {len(unique_msgs)}건" +
                                 (f" (중복 {skipped_msgs}건 건너뜀)" if skipped_msgs else ""))
                    if defaulted_term_count:
                        parts.append(f"학기 미지정 {defaulted_term_count}건 → '{DEFAULT_TERM}' 자동 배정")
                    flash("추가 완료: " + " · ".join(parts), "success")
                else:
                    save_data(new_students, new_records, new_homeworks, new_messages)
                    tail = ""
                    if defaulted_term_count:
                        tail = f" (학기 미지정 {defaulted_term_count}건은 '{DEFAULT_TERM}'으로 자동 배정)"
                    flash(
                        f"덮어쓰기 완료: 학생 {len(new_students)}명, 기록 {len(new_records)}건, "
                        f"숙제 {len(new_homeworks)}건, 한마디 {len(new_messages)}건으로 전체 교체됨.{tail}",
                        "success"
                    )
        return redirect(url_for("admin"))

    data = load_data()
    return render_template(
        "admin.html",
        students=data["students"],
        record_count=len(data["records"]),
        last_update=datetime.fromtimestamp(data["mtime"]).strftime("%Y-%m-%d %H:%M") if data["mtime"] else "—",
        excel_exists=os.path.exists(EXCEL_PATH),
    )


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


# ─── 라우트: 관리자 — 기록 관리 ──────────────────────────────
@app.route("/admin/records", methods=["GET", "POST"])
@admin_required
def admin_records():
    data = load_data()

    if request.method == "POST":
        action = request.form.get("action", "add")

        if action == "bulk_delete":
            # 선택된 인덱스들을 받아서 일괄 삭제
            raw_indices = request.form.getlist("selected")
            try:
                indices = sorted({int(i) for i in raw_indices}, reverse=True)
            except (ValueError, TypeError):
                indices = []

            if not indices:
                flash("삭제할 기록을 선택해주세요.", "error")
            else:
                removed = 0
                for i in indices:
                    if 0 <= i < len(data["records"]):
                        del data["records"][i]
                        removed += 1
                save_data(data["students"], data["records"])
                flash(f"{removed}건의 기록이 일괄 삭제되었습니다.", "success")
            return redirect(url_for("admin_records",
                                    student=request.args.get("student", ""),
                                    category=request.args.get("category", "")))

        if action == "bulk_save":
            # 인라인 편집된 점수/피드백/비고/완료를 일괄 저장
            changed = 0
            for i in range(len(data["records"])):
                # 폼에 present_{i}=1 마커가 없으면 (필터로 가려진 행) 건너뜀
                if request.form.get(f"present_{i}") != "1":
                    continue
                r = data["records"][i]
                new_score = request.form.get(f"score_{i}", "").strip()
                new_feedback = request.form.get(f"feedback_{i}", "").strip()
                new_flag = request.form.get(f"flag_{i}", "").strip()
                new_resolved = request.form.get(f"resolved_{i}") == "on"
                if (new_score != r.get("score", "") or
                    new_feedback != r.get("feedback", "") or
                    new_flag != r.get("flag", "") or
                    new_resolved != bool(r.get("resolved"))):
                    r["score"] = new_score
                    r["feedback"] = new_feedback
                    r["flag"] = new_flag
                    r["resolved"] = new_resolved
                    changed += 1
            if changed > 0:
                save_data(data["students"], data["records"])
                flash(f"{changed}건의 기록이 일괄 수정되었습니다.", "success")
            else:
                flash("변경된 내용이 없습니다.", "success")
            return redirect(url_for("admin_records",
                                    student=request.args.get("student", ""),
                                    category=request.args.get("category", "")))

        # 기본: 새 기록 추가
        date = request.form.get("date", "").strip()
        student_code = request.form.get("student_code", "").strip()
        category = request.form.get("category", "").strip()
        score = request.form.get("score", "").strip()
        feedback = request.form.get("feedback", "").strip()
        flag = request.form.get("flag", "").strip()
        resolved = request.form.get("resolved") == "on"
        term = request.form.get("term", "").strip() or DEFAULT_TERM

        if not date:
            flash("날짜를 입력해주세요.", "error")
        elif not student_code:
            flash("학생을 선택해주세요.", "error")
        elif student_code not in data["students"]:
            flash("등록되지 않은 학생코드입니다.", "error")
        else:
            data["records"].append({
                "date": date,
                "student_code": student_code,
                "category": category or "기타",
                "score": score,
                "feedback": feedback,
                "flag": flag,
                "resolved": resolved,
                "term": term,
            })
            save_data(data["students"], data["records"])
            flash(f"{data['students'][student_code]['name']} 학생의 기록이 추가되었습니다.", "success")
        return redirect(url_for("admin_records",
                                student=request.args.get("student", ""),
                                category=request.args.get("category", "")))

    filter_student = request.args.get("student", "").strip()
    filter_category = request.args.get("category", "").strip()
    filter_term = request.args.get("term", "").strip()

    indexed = list(enumerate(data["records"]))
    if filter_student:
        indexed = [(i, r) for i, r in indexed if r["student_code"] == filter_student]
    if filter_category:
        indexed = [(i, r) for i, r in indexed if r["category"] == filter_category]
    if filter_term:
        indexed = [(i, r) for i, r in indexed if (r.get("term") or DEFAULT_TERM) == filter_term]

    indexed.sort(key=lambda x: x[1]["date"], reverse=True)

    categories = sorted({r["category"] for r in data["records"] if r["category"]})
    today = datetime.now().strftime("%Y-%m-%d")

    return render_template(
        "admin_records.html",
        records=indexed,
        students=data["students"],
        categories=categories,
        filter_student=filter_student,
        filter_category=filter_category,
        filter_term=filter_term,
        today=today,
        terms=TERMS,
        default_term=DEFAULT_TERM,
    )


@app.route("/admin/records/<int:idx>/edit", methods=["GET", "POST"])
@admin_required
def admin_record_edit(idx):
    data = load_data()

    if idx < 0 or idx >= len(data["records"]):
        flash("기록을 찾을 수 없습니다.", "error")
        return redirect(url_for("admin_records"))

    if request.method == "POST":
        if request.form.get("action") == "delete":
            del data["records"][idx]
            save_data(data["students"], data["records"])
            flash("기록이 삭제되었습니다.", "success")
        else:
            data["records"][idx] = {
                "date": request.form.get("date", "").strip(),
                "student_code": request.form.get("student_code", "").strip(),
                "category": request.form.get("category", "").strip() or "기타",
                "score": request.form.get("score", "").strip(),
                "feedback": request.form.get("feedback", "").strip(),
                "flag": request.form.get("flag", "").strip(),
                "resolved": request.form.get("resolved") == "on",
                "term": request.form.get("term", "").strip() or DEFAULT_TERM,
            }
            save_data(data["students"], data["records"])
            flash("기록이 수정되었습니다.", "success")
        return redirect(url_for("admin_records"))

    return render_template(
        "admin_record_edit.html",
        record=data["records"][idx],
        idx=idx,
        students=data["students"],
        categories=sorted({r["category"] for r in data["records"] if r["category"]}),
        terms=TERMS,
        default_term=DEFAULT_TERM,
    )


# ─── 라우트: 관리자 — 학생 관리 ──────────────────────────────
@app.route("/admin/students", methods=["GET", "POST"])
@admin_required
def admin_students():
    data = load_data()

    if request.method == "POST":
        code = request.form.get("student_code", "").strip()
        name = request.form.get("student_name", "").strip()
        pin = request.form.get("pin", "").strip()
        parent = request.form.get("parent", "").strip()

        if not code or not name:
            flash("학생코드와 이름은 필수입니다.", "error")
        elif code in data["students"]:
            flash(f"학생코드 '{code}'는 이미 사용 중입니다.", "error")
        else:
            data["students"][code] = {"name": name, "pin": pin, "parent": parent}
            save_data(data["students"], data["records"])
            flash(f"학생 '{name}'이(가) 추가되었습니다.", "success")
        return redirect(url_for("admin_students"))

    return render_template("admin_students.html", students=data["students"])


@app.route("/admin/students/<code>/edit", methods=["GET", "POST"])
@admin_required
def admin_student_edit(code):
    data = load_data()

    if code not in data["students"]:
        flash("학생을 찾을 수 없습니다.", "error")
        return redirect(url_for("admin_students"))

    if request.method == "POST":
        if request.form.get("action") == "delete":
            del data["students"][code]
            data["records"] = [r for r in data["records"] if r["student_code"] != code]
            data["messages"] = [m for m in data["messages"] if m.get("student_code") != code]
            save_data(data["students"], data["records"], data["homework"], data["messages"])
            flash(f"학생 '{code}'와(과) 관련된 모든 기록·한마디가 삭제되었습니다.", "success")
            return redirect(url_for("admin_students"))

        new_code = request.form.get("student_code", "").strip()
        new_name = request.form.get("student_name", "").strip()
        new_pin = request.form.get("pin", "").strip()
        new_parent = request.form.get("parent", "").strip()

        if not new_code or not new_name:
            flash("학생코드와 이름은 필수입니다.", "error")
            return redirect(url_for("admin_student_edit", code=code))

        if new_code != code and new_code in data["students"]:
            flash(f"학생코드 '{new_code}'는 이미 사용 중입니다.", "error")
            return redirect(url_for("admin_student_edit", code=code))

        if new_code != code:
            data["students"].pop(code)
            for r in data["records"]:
                if r["student_code"] == code:
                    r["student_code"] = new_code

        data["students"][new_code] = {"name": new_name, "pin": new_pin, "parent": new_parent}
        save_data(data["students"], data["records"])
        flash("학생 정보가 수정되었습니다.", "success")
        return redirect(url_for("admin_students"))

    return render_template(
        "admin_student_edit.html",
        student=data["students"][code],
        code=code,
    )


# ─── 라우트: 관리자 — 클리닉 관리 ────────────────────────
def _flash_clinic_result(date_str, added, updated, skipped):
    parts = [f"{date_str} 클리닉 기록"]
    if added:
        parts.append(f"신규 추가 {len(added)}명 ({', '.join(added)})")
    if updated:
        parts.append(f"메모 추가됨 {len(updated)}명 ({', '.join(updated)})")
    if skipped:
        parts.append(f"이미 있어 건너뜀 {len(skipped)}명")
    if not (added or updated or skipped):
        parts.append("처리할 학생이 없습니다")
    flash(" · ".join(parts), "success")


@app.route("/admin/clinic", methods=["GET", "POST"])
@admin_required
def admin_clinic():
    if request.method == "POST":
        action = request.form.get("action", "")

        # 1) 개별 클리닉 수동 추가 (학생 복수 선택 + 날짜 + 자동피드백 여부 + 메모)
        if action == "add_manual":
            date_str = request.form.get("date", "").strip()
            codes = request.form.getlist("student_codes")
            auto_fb = request.form.get("auto_feedback") == "on"
            note = request.form.get("note", "").strip()
            if not date_str:
                flash("날짜를 선택해주세요.", "error")
            elif not codes:
                flash("학생을 한 명 이상 선택해주세요.", "error")
            else:
                added, updated, skipped = add_clinic_records(
                    date_str, codes, auto_feedback=auto_fb, note=note)
                _flash_clinic_result(date_str, added, updated, skipped)
            return redirect(url_for("admin_clinic"))

        # 2) 요일별 정기 명단으로 한 번에 추가 (수동 트리거, 자동 스케줄 아님)
        if action == "add_roster":
            try:
                weekday = int(request.form.get("weekday", -1))
            except (ValueError, TypeError):
                weekday = -1
            date_str = request.form.get("date", "").strip()
            names = load_clinic_roster().get(str(weekday), [])
            if weekday not in (1, 3, 5) or not date_str:
                flash("올바르지 않은 요청입니다.", "error")
            elif not names:
                flash("해당 요일 정기 명단이 비어 있습니다. 먼저 명단을 저장하세요.", "error")
            else:
                codes = _resolve_names_to_codes(names)
                added, updated, skipped = add_clinic_records(date_str, codes, auto_feedback=True)
                _flash_clinic_result(date_str, added, updated, skipped)
            return redirect(url_for("admin_clinic"))

        # 3) 요일별 정기 명단 저장
        if action == "save_roster":
            roster = {}
            for wd, _label in CLINIC_WEEKDAYS:
                raw = request.form.get(f"roster_{wd}", "")
                names = [x.strip() for x in raw.replace(",", "\n").splitlines() if x.strip()]
                roster[str(wd)] = names
            save_clinic_roster(roster)
            flash("요일별 정기 명단을 저장했습니다.", "success")
            return redirect(url_for("admin_clinic"))

        # 4) 클리닉 기록 삭제
        if action == "delete":
            try:
                idx = int(request.form.get("idx", -1))
            except (ValueError, TypeError):
                idx = -1
            data = load_data()
            if (0 <= idx < len(data["records"])
                    and data["records"][idx].get("category") == CLINIC_CATEGORY):
                removed = data["records"].pop(idx)
                nm = data["students"].get(removed.get("student_code"), {}).get("name", "?")
                save_data(data["students"], data["records"])
                flash(f"{removed.get('date')} · {nm} 클리닉 기록을 삭제했습니다.", "success")
            else:
                flash("삭제할 기록을 찾지 못했습니다. (목록이 바뀌었을 수 있어요)", "error")
            return redirect(url_for("admin_clinic"))

        flash("올바르지 않은 요청입니다.", "error")
        return redirect(url_for("admin_clinic"))

    # GET
    data = load_data()
    recent_clinic = sorted(
        [(i, r) for i, r in enumerate(data["records"]) if r.get("category") == CLINIC_CATEGORY],
        key=lambda t: t[1].get("date", ""),
        reverse=True,
    )[:30]
    students_sorted = sorted(data["students"].items(), key=lambda kv: kv[1].get("name", ""))
    roster = load_clinic_roster()
    weekday_rows = [
        {"num": wd, "label": label, "names": roster.get(str(wd), []),
         "this_date": _this_week_weekday_date(wd)}
        for wd, label in CLINIC_WEEKDAYS
    ]
    return render_template(
        "admin_clinic.html",
        students_sorted=students_sorted,
        recent=recent_clinic,
        weekday_rows=weekday_rows,
        today=datetime.now().strftime("%Y-%m-%d"),
        students=data["students"],
    )


# ─── 라우트: 관리자 — 시험 등수 모아보기 ──────────────────
@app.route("/admin/rankings")
@admin_required
def admin_rankings():
    data = load_data()

    # 학기 필터 (기본: 2학기 정규반)
    selected_term = request.args.get("term", DEFAULT_TERM)
    if selected_term not in TERMS:
        selected_term = DEFAULT_TERM

    def _term_of(r):
        return r.get("term") or DEFAULT_TERM

    # 학기별 카운트 (탭 뱃지용)
    term_counts = {t: 0 for t in TERMS}
    for r in data["records"]:
        t = _term_of(r)
        if t in term_counts:
            term_counts[t] += 1

    # 항목별로 학생별 점수 모으기 (숫자 점수만, 선택된 학기만)
    cat_scores = defaultdict(lambda: defaultdict(list))  # cat -> {code -> [scores]}
    for r in data["records"]:
        if _term_of(r) != selected_term:
            continue
        sc = _parse_score(r["score"])
        if sc is None:
            continue
        cat = r["category"] or "기타"
        cat_scores[cat][r["student_code"]].append(sc)

    # 항목별 등수 계산 (경쟁식: 1, 2, 2, 4)
    rankings = {}
    for cat, code_scores in cat_scores.items():
        entries = []
        for code, scores in code_scores.items():
            entries.append({
                "code": code,
                "name": data["students"].get(code, {}).get("name", "?"),
                "best": max(scores),
                "avg": round(sum(scores) / len(scores), 1),
                "count": len(scores),
            })
        # 최고점 기준 내림차순
        entries.sort(key=lambda e: (-e["best"], -e["avg"], e["name"]))
        prev_best = None
        current_rank = 0
        for i, e in enumerate(entries):
            if e["best"] != prev_best:
                current_rank = i + 1
                prev_best = e["best"]
            e["rank"] = current_rank
        rankings[cat] = entries

    # 항목명 알파벳 정렬 (한국어 가나다)
    sorted_cats = sorted(rankings.keys())

    # 종합 순위: 메달 개수 기준 (시험마다 만점 다르므로 평균 의미 없음)
    # 금=1등, 은=2등, 동=3등
    medal_counts = defaultdict(lambda: {"gold": 0, "silver": 0, "bronze": 0, "categories": 0})
    for cat, entries in rankings.items():
        for e in entries:
            code = e["code"]
            medal_counts[code]["categories"] += 1
            if e["rank"] == 1:
                medal_counts[code]["gold"] += 1
            elif e["rank"] == 2:
                medal_counts[code]["silver"] += 1
            elif e["rank"] == 3:
                medal_counts[code]["bronze"] += 1

    overall = []
    for code, counts in medal_counts.items():
        overall.append({
            "code": code,
            "name": data["students"].get(code, {}).get("name", "?"),
            "gold": counts["gold"],
            "silver": counts["silver"],
            "bronze": counts["bronze"],
            "total_medals": counts["gold"] + counts["silver"] + counts["bronze"],
            "category_count": counts["categories"],
        })

    # 정렬: 금 많은 순 → 은 많은 순 → 동 많은 순 → 이름
    overall.sort(key=lambda x: (-x["gold"], -x["silver"], -x["bronze"], x["name"]))

    # 경쟁식 등수 (금·은·동 개수 모두 같으면 같은 등수)
    prev_key = None
    current_rank = 0
    for i, e in enumerate(overall):
        key = (e["gold"], e["silver"], e["bronze"])
        if key != prev_key:
            current_rank = i + 1
            prev_key = key
        e["rank"] = current_rank

    return render_template(
        "admin_rankings.html",
        rankings=rankings,
        categories=sorted_cats,
        overall=overall,
        total_students=len(data["students"]),
        terms=TERMS,
        selected_term=selected_term,
        term_counts=term_counts,
        default_term=DEFAULT_TERM,
    )


# ─── 라우트: 관리자 — 특이사항 (재시/숙제미비/추가과제) ────
@app.route("/admin/flags", methods=["GET", "POST"])
@admin_required
def admin_flags():
    data = load_data()

    if request.method == "POST":
        action = request.form.get("action", "")
        try:
            idx = int(request.form.get("idx", -1))
        except (ValueError, TypeError):
            idx = -1

        if 0 <= idx < len(data["records"]):
            if action == "resolve":
                data["records"][idx]["resolved"] = True
                save_data(data["students"], data["records"])
                flash("완료 처리되었습니다.", "success")
            elif action == "unresolve":
                data["records"][idx]["resolved"] = False
                save_data(data["students"], data["records"])
                flash("미처리로 되돌렸습니다.", "success")
            elif action == "clear_flag":
                data["records"][idx]["flag"] = ""
                data["records"][idx]["resolved"] = False
                save_data(data["students"], data["records"])
                flash("비고가 제거되었습니다.", "success")
        return redirect(url_for("admin_flags"))

    # 미처리 특이사항: flag 있고 resolved 아닌 것
    active = [(i, r) for i, r in enumerate(data["records"])
              if r.get("flag") and not r.get("resolved")]
    active_by_date = defaultdict(list)
    for i, r in active:
        active_by_date[r["date"]].append((i, r))
    active_list = sorted(active_by_date.items(), key=lambda x: x[0], reverse=True)

    # 처리 완료 이력 (최근 30건)
    resolved_recs = [(i, r) for i, r in enumerate(data["records"])
                     if r.get("flag") and r.get("resolved")]
    resolved_recs.sort(key=lambda x: x[1]["date"], reverse=True)
    resolved_recs = resolved_recs[:30]

    # 플래그 종류별 카운트
    flag_counts = defaultdict(int)
    for _, r in active:
        flag_counts[r["flag"]] += 1

    return render_template(
        "admin_flags.html",
        active_by_date=active_list,
        resolved_recs=resolved_recs,
        students=data["students"],
        total_active=len(active),
        flag_counts=dict(flag_counts),
    )


# ─── 라우트: 관리자 — 공지사항(이번주 숙제 + 한마디) ────────
@app.route("/admin/announcements", methods=["GET", "POST"])
@admin_required
def admin_announcements():
    data = load_data()

    if request.method == "POST":
        action = request.form.get("action", "")

        # ─── 이번주 숙제 ────────────────────────────────
        if action == "add_homework":
            content = request.form.get("homework", "").strip()
            if not content:
                flash("숙제 내용을 입력해주세요.", "error")
            else:
                new_hws = list(data["homeworks"]) + [{"content": content, "published": True}]
                save_data(homeworks=new_hws)
                flash("새 숙제가 게시되었습니다.", "success")

        elif action in ("publish_homework", "unpublish_homework", "delete_homework"):
            try:
                idx = int(request.form.get("idx", -1))
            except ValueError:
                idx = -1
            if 0 <= idx < len(data["homeworks"]):
                new_hws = list(data["homeworks"])
                if action == "delete_homework":
                    del new_hws[idx]
                    flash("숙제가 삭제되었습니다.", "success")
                elif action == "publish_homework":
                    new_hws[idx] = dict(new_hws[idx], published=True)
                    flash("숙제가 다시 게시되었습니다.", "success")
                else:
                    new_hws[idx] = dict(new_hws[idx], published=False)
                    flash("숙제를 내렸습니다 (보관됨).", "success")
                save_data(homeworks=new_hws)
            else:
                flash("대상 숙제를 찾을 수 없습니다.", "error")

        # ─── 한마디 ─────────────────────────────────────
        elif action == "add_message":
            content = request.form.get("content", "").strip()
            student_code = request.form.get("student_code", "").strip()
            if not content:
                flash("한마디 내용을 입력해주세요.", "error")
            elif student_code and student_code not in data["students"]:
                flash("선택한 학생코드가 존재하지 않습니다.", "error")
            else:
                new_messages = list(data["messages"]) + [{
                    "student_code": student_code,
                    "content": content,
                    "published": True,
                }]
                save_data(messages=new_messages)
                target = data["students"][student_code]["name"] if student_code else "전체"
                flash(f"한마디가 게시되었습니다. (대상: {target})", "success")

        elif action in ("publish_message", "unpublish_message", "delete_message"):
            try:
                idx = int(request.form.get("idx", -1))
            except ValueError:
                idx = -1
            if 0 <= idx < len(data["messages"]):
                new_messages = list(data["messages"])
                if action == "delete_message":
                    del new_messages[idx]
                    flash("한마디가 삭제되었습니다.", "success")
                elif action == "publish_message":
                    new_messages[idx] = dict(new_messages[idx], published=True)
                    flash("한마디가 다시 게시되었습니다.", "success")
                else:
                    new_messages[idx] = dict(new_messages[idx], published=False)
                    flash("한마디를 내렸습니다 (보관됨).", "success")
                save_data(messages=new_messages)
            else:
                flash("대상 한마디를 찾을 수 없습니다.", "error")

        return redirect(url_for("admin_announcements"))

    # 분리 (인덱스 유지 위해 enumerate)
    hws_indexed = list(enumerate(data["homeworks"]))
    hws_published = [(i, h) for i, h in hws_indexed if h.get("published")]
    hws_archived = [(i, h) for i, h in hws_indexed if not h.get("published")]

    msgs_indexed = list(enumerate(data["messages"]))
    msgs_published = [(i, m) for i, m in msgs_indexed if m.get("published")]
    msgs_archived = [(i, m) for i, m in msgs_indexed if not m.get("published")]

    return render_template(
        "admin_announcements.html",
        hws_published=hws_published,
        hws_archived=hws_archived,
        msgs_published=msgs_published,
        msgs_archived=msgs_archived,
        students=data["students"],
    )


# ─── 라우트: Excel 다운로드/템플릿 ───────────────────────────
@app.route("/admin/download")
@admin_required
def admin_download():
    """현재 데이터를 Excel 파일로 다운로드.
    다운로드 직전에 최신 형식(비고/완료 컬럼 포함)으로 재저장하여 항상 신형식 보장.
    """
    if not os.path.exists(EXCEL_PATH):
        flash("아직 데이터가 없습니다. 빈 템플릿을 받으시려면 'Excel 템플릿 다운로드'를 이용하세요.", "error")
        return redirect(url_for("admin"))
    # 신형식 보장을 위해 재저장
    data = load_data()
    save_data(data["students"], data["records"], data["homeworks"], data["messages"])
    return send_file(
        EXCEL_PATH,
        as_attachment=True,
        download_name=f"students_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/template")
def download_template():
    """빈 Excel 템플릿 다운로드 (3시트: 학생명단, 기록, 공지사항)."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "학생명단"
    ws1.append(["학생코드", "학생이름", "PIN", "학부모이름(선택)"])
    ws1.append(["S001", "김민지", "1234", "김민지 어머니"])
    ws1.append(["S002", "이도윤", "5678", "이도윤 어머니"])

    ws2 = wb.create_sheet("기록")
    ws2.append(["날짜", "학생코드", "학생이름", "항목", "점수", "피드백", "비고", "완료", "학기/과정"])
    ws2.append(["2026-05-18", "S001", "김민지", "Daily Test", "92",
                "어휘 문제에서 실수가 있었지만 독해는 완벽했습니다.", "", "", "2학기 정규반"])
    ws2.append(["2026-05-18", "S001", "김민지", "숙제", "완료", "꼼꼼하게 잘 했습니다.", "", "", "2학기 정규반"])
    ws2.append(["2026-05-18", "S002", "이도윤", "단어시험", "60",
                "기준 미달, 재시 필요.", "재시", "", "2학기 정규반"])
    ws2.append(["2026-01-15", "S002", "이도윤", "중간평가", "38",
                "재시 대상.", "재시", "", "방학특강"])

    ws3 = wb.create_sheet("공지사항")
    ws3.append(["종류", "대상학생코드", "학생이름", "내용", "상태"])
    ws3.append(["이번주숙제", "", "", "워크북 32-45쪽 풀고, 단어 50개 외워오기", "게시"])
    ws3.append(["이번주숙제", "", "", "지난주 숙제: 모의고사 1회 풀어오기", "내림"])
    ws3.append(["한마디", "", "", "시험 기간 화이팅! 모두 잘 할 수 있을거예요.", "게시"])
    ws3.append(["한마디", "S001", "김민지", "단어시험 1등 축하해요. 다음주도 기대할게요!", "게시"])

    for ws, widths in [
        (ws1, [10, 12, 8, 18]),
        (ws2, [12, 10, 12, 14, 8, 40, 12, 8, 14]),
        (ws3, [14, 14, 12, 50, 10]),
    ]:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    _add_records_dropdowns(ws2)
    _add_term_dropdown(ws2)
    _add_announcements_dropdown(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="students_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── 헬스체크 ────────────────────────────────────────────────
@app.route("/healthz")
def healthz():
    return "ok", 200


# ─── 1회성 마이그레이션: 기존 기록 → 방학특강 표시 ───────────
def run_one_time_term_migration():
    """1회성: 기존 모든 기록의 학기/과정 필드가 비어있으면 '방학특강'으로 표시.
    이후 신규 기록은 폼 기본값(2학기 정규반)으로 저장됨.
    """
    import json
    if os.path.exists(_TERM_MIGRATION_MARKER):
        return
    try:
        data = load_data()
        records = list(data["records"])
        modified = 0
        for r in records:
            if not r.get("term"):
                r["term"] = "방학특강"
                modified += 1
        if modified > 0:
            save_data(data["students"], records)
        os.makedirs(os.path.dirname(_TERM_MIGRATION_MARKER), exist_ok=True)
        with open(_TERM_MIGRATION_MARKER, "w", encoding="utf-8") as f:
            json.dump({
                "ran_at": datetime.now().isoformat(),
                "modified_count": modified,
            }, f, ensure_ascii=False)
        print(f"[INFO] Term migration completed: {modified} records marked as 방학특강")
    except Exception as e:
        print(f"[WARN] Term migration failed: {e}")


# 앱 import 시점에 마이그레이션 실행
run_one_time_term_migration()


# ─── 메인 ────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
